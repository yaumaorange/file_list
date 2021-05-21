import os
import openpyxl

def wirte_excel(file_name, sheet):
	c = 0
	for f in file_name:
		if '已開標' not in f:
			c += 1
			sheet.cell(row=c, column=1, value=f[8:])



n_bid = os.listdir('C:\\Users\\TD100-E580\\Documents\\21_標案資料\\標案資料')
bid_open = os.listdir('C:\\Users\\TD100-E580\\Documents\\21_標案資料\\標案資料\\已開標\\110年度')
print(bid_open)

wb = openpyxl.Workbook()
sheet2 = wb.create_sheet('未開標')
sheet1 = wb.create_sheet('已開標')

wirte_excel(bid_open, sheet1)
wirte_excel(n_bid, sheet2)
wb.save('資料夾列表.xlsx')
