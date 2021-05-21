import os
import openpyxl
import datetime

#讀取excel
def read_excel(sheet, bid):
	for cell in sheet['A']:
		bid.append(cell.value)

#讀取資料夾名稱
def read_filename(file_list, bid):
	for fl in file_list:
		if fl[8:] in bid:
			continue
		bid.append(fl[8:])

def wirte_nb(file_name, sheet):
	c = 1
	for f in file_name:
		if '已開標' not in f:
			c += 1
			sheet.cell(row=c, column=1, value=f[8:])

def wirte_ob(file_name, sheet):
	for f in file_name:
		c = 1
		if '案件名稱' not in f:
			c += 1
			sheet.cell(row=c, column=1, value=f)


wb = openpyxl.load_workbook('資料夾列表.xlsx')

sheet1 = wb['未開標']
sheet2 = wb['已開標']

not_bid = os.listdir('C:\\Users\\TD100-E580\\Documents\\21_標案資料\\標案資料')
open_bid = os.listdir('C:\\Users\\TD100-E580\\Documents\\21_標案資料\\標案資料\\已開標\\110年度')

bid = []
read_excel(sheet2, bid)
read_filename(open_bid, bid)
wirte_nb(not_bid, sheet1)
wirte_ob(bid, sheet2)
today = str(datetime.date.today())

wb.save('標案清單' + today + '.xlsx')
